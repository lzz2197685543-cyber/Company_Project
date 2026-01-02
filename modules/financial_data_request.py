import requests
from utils.cookie_manager_financial import CookieManager
from utils.logger import get_logger
from pathlib import Path
from datetime import datetime, timezone, timedelta
import asyncio
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

"""用request获取temu财务明细"""

class Temu_Financial_Data:
    def __init__(self, shop_name, month_str):
        self.month_str = month_str
        self.shop_name = shop_name

        self.headers = {
            'origin': 'https://seller.kuajingmaihuo.com',
            'referer': 'https://seller.kuajingmaihuo.com/labor/bill',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        }

        self.url = 'https://seller.kuajingmaihuo.com/api/merchant/fund/detail/pageSearch'
        self.cookie_manager = CookieManager(shop_name)
        self.logger = get_logger("financial_data")

    def get_month_date_range(self, month_str: str) -> dict:
        year, month = map(int, month_str.split("-"))

        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year + 1, 1, 1) - timedelta(seconds=1)
        else:
            end = datetime(year, month + 1, 1) - timedelta(seconds=1)

        return {
            "start_date": start.strftime("%Y-%m-%d"),
            "end_date": end.strftime("%Y-%m-%d"),

        }

    # ======================
    # 时间转换
    # ======================
    def date_range_to_timestamp_ms(self, tz_offset=8):
        start_time=self.get_month_date_range(self.month_str)['start_date']
        end_time=self.get_month_date_range(self.month_str)['end_date']

        tz = timezone(timedelta(hours=tz_offset))

        start_dt = datetime.strptime(start_time, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, tzinfo=tz
        )
        end_dt = datetime.strptime(end_time, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, tzinfo=tz
        )

        return int(start_dt.timestamp() * 1000), int(end_dt.timestamp() * 1000)

    # ======================
    # 请求接口（只负责请求）
    # ======================
    def get_info(self, page, cookies, shop_id):
        self.headers["mallid"] = str(shop_id)
        begin_ts, end_ts = self.date_range_to_timestamp_ms()

        payload = {
            "beginTime": begin_ts,
            "endTime": end_ts,
            "pageSize": 100,
            "pageNum": page,
        }

        response = requests.post(
            self.url,
            headers=self.headers,
            cookies=cookies,
            json=payload,
            timeout=15
        )
        print(response.text[:200])

        response.raise_for_status()
        return response.json()

    # ======================
    # 解析数据
    # ======================
    def parse_data(self, json_data):
        try:
            result_list = json_data["result"]["resultList"]
        except (KeyError, TypeError):
            self.logger.error("接口数据结构异常")
            return []


        items = []
        for i in result_list:
            money_type = i.get("moneyChangeType")
            sign = "+" if money_type == 1 else "-" if money_type == 2 else ""

            # 取金额数字，防止 None
            amount_str = i.get("amount", "")
            amount_value = amount_str.split("¥")[1] if "¥" in amount_str else amount_str

            items.append({
                "财务时间": i.get("transactionTime"),
                "财务类型": i.get("fundTypeDesc"),
                "币种": i.get("currencyType"),
                "收支金额": f"{sign}{amount_value}",
                "备注": i.get("remark", "")
            })

        return items

    # ======================
    # 保存 xlsx
    # ======================

    def save_items(self, items):
        if not items:
            return

        out_dir = Path(__file__).resolve().parent.parent / "data" / "financial" / "temu"
        out_dir.mkdir(parents=True, exist_ok=True)

        fname = out_dir / f"{self.shop_name}_{self.month_str.split('-')[1]}_对账单.xlsx"

        if fname.exists():
            wb = load_workbook(fname)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            headers = list(items[0].keys())
            ws.append(headers)

        headers = [cell.value for cell in ws[1]]

        for item in items:
            ws.append([item.get(h, "") for h in headers])

        # 自动列宽
        for col_idx, col_name in enumerate(headers, 1):
            max_length = len(str(col_name))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        wb.save(fname)

    # ======================
    # 主流程
    # ======================
    async def run(self):
        self.logger.info(f"开始爬取店铺：{self.shop_name}")
        page = 1

        while True:
            json_data = None

            for attempt in range(3):
                try:
                    cookies, shop_id = await self.cookie_manager.get_auth()
                    json_data = self.get_info(page, cookies, shop_id)

                    # 登录态判断
                    if json_data.get("error_code") == 40001 or \
                       json_data.get("error_msg") == "登录过期，请重新登录":
                        raise PermissionError("登录过期")

                    break  # 成功跳出 retry

                except PermissionError:
                    self.logger.warning(f"[{self.shop_name}] 登录失效，刷新 cookie（第 {attempt+1} 次）")
                    await self.cookie_manager.refresh()
                    await asyncio.sleep(2)

                except Exception as e:
                    self.logger.error(f"[{self.shop_name}] 第 {attempt+1} 次请求失败: {e}")
                    await self.cookie_manager.refresh()
                    await asyncio.sleep(2)

            if not json_data:
                self.logger.error(f"[{self.shop_name}] 重试失败，终止任务")
                return

            items = self.parse_data(json_data)
            self.logger.info(f"第 {page} 页返回数量: {len(json_data['result']['resultList'])}")
            self.save_items(items)

            self.logger.info(f"店铺-{self.shop_name}-第 {page} 页保存成功")

            if len(items) < 100:
                self.logger.info("已到最后一页")
                break

            page += 1

#
# if __name__ == '__main__':
#     month_str='2025-12'
#     t=Temu_Financial_Data("102-Temu全托管",month_str)
#     asyncio.run(t.run())


