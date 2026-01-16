import requests
from utils.cookie_manager import CookieManager
from utils.logger import get_logger
from pathlib import Path
from datetime import datetime, timezone, timedelta
import asyncio
import csv
from utils.dingtalk_bot import ding_bot_send
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

"""shein提现"""


class Shein_Financial_Data_Tixian:
    def __init__(self, shop_name, month_str):
        self.month_str = month_str
        self.shop_name = shop_name
        self.cookie_manager = CookieManager(shop_name)
        self.cookies = None

        self.headers = {
            'origin': 'https://sso.geiwohuo.com',
            'origin-url': 'https://sso.geiwohuo.com',
            'priority': 'u=1, i',
            'referer': 'https://sso.geiwohuo.com/',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        }

        self.url = 'https://sso.geiwohuo.com/mws/mwms/sso/withdraw/transferRecordList'

        self.logger = get_logger("financial_data_tixian")

    def is_cookie_invalid(self, json_data):
        """
        统一判断 cookie 是否失效
        """
        # 请求异常
        if not json_data:
            return True

        # get_info 主动标记
        if json_data.get("msg") == "20302":
            return True

        if not isinstance(json_data, dict):
            return True

        return False

    def get_month_date_range(self, month_str: str) -> dict:
        year, month = map(int, month_str.split("-"))

        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year + 1, 1, 1) - timedelta(seconds=1)
        else:
            end = datetime(year, month + 1, 1) - timedelta(seconds=1)

        return {
            "start_date": start.strftime("%Y-%m-%d"),
            "end_date": end.strftime("%Y-%m-%d"),

        }

    # ======================
    # 时间转换
    # ======================
    def date_range_to_timestamp_ms(self, tz_offset=8):
        start_time = self.get_month_date_range(self.month_str)["start_date"]
        end_time = self.get_month_date_range(self.month_str)["end_date"]

        tz = timezone(timedelta(hours=tz_offset))

        start_dt = datetime.strptime(start_time, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, tzinfo=tz
        )
        end_dt = datetime.strptime(end_time, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, tzinfo=tz
        )

        return int(start_dt.timestamp() * 1000), int(end_dt.timestamp() * 1000)

    def format_ts(self, ts):
        """
        时间戳转：2025-12-30 22:06:26
        支持 秒 / 毫秒
        """
        if not ts:
            return None

        ts = int(ts)

        # 毫秒时间戳
        if ts > 1e12:
            ts = ts / 1000

        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")

    def get_info(self, page, cookies):
        self.cookies = cookies
        begin_ts, end_ts = self.date_range_to_timestamp_ms()
        json_data = {
            'reqSystemCode': 'mws-front',
            # 'supplierId': 4120251,
            'pageNum': page,
            'pageSize': 100,
            'createTimeStart': begin_ts,
            'createTimeEnd': end_ts,
        }
        try:
            response = requests.post(
                url=self.url,
                cookies=self.cookies,
                headers=self.headers,
                json=json_data,
            )
            print(response.text[:200])
            return response.json()
        except Exception as e:
            self.logger.error(f'请求响应失败:{e}')

    def parse_data(self, json_data):
        try:
            result_list = json_data['info']['list']
        except (KeyError, TypeError):
            self.logger.error("接口数据结构异常")
            return []
        items = []
        for i in result_list:
            items.append({
                "交易单号": i.get('withdrawNo'),
                "提现时间": self.format_ts(i.get('createTime')),
                "提现成功时间": self.format_ts(i.get('transferSuccessTime')),
                "更新时间": self.format_ts(i.get('lastUpdateTime')),
                "提现明细单号": i.get('transferNo'),
                "收款账户": i.get('sourceAccountValue'),
                "收款账户所在地": i.get('accountAreaCode'),
                "净金额": i.get('netAmount'),
                "净金额币种": i.get('currency'),
                "保证金": i.get('depositAmount'),
                "保证金币种": i.get('currency'),
                "手续费": i.get('commissionAmount'),
                "手续费币种": i.get('currency'),
                "汇率": i.get('exchangeRate'),
                "收款金额": i.get('receivingAmount'),
                "收款币种": i.get('receivingCurrency'),
                "提现状态": i.get('withdrawStatusDesc'),
                "失败原因": i.get('failReason', '')
            })

        return items

    def save_batch(self, items):
        if not items:
            return

        out_dir = Path(__file__).resolve().parent.parent / "data" / "financial" / (str(self.month_str.split('-')[1]) + '月份') / "shein"
        out_dir.mkdir(parents=True, exist_ok=True)


        fname = out_dir / f"{self.shop_name}_{self.month_str.split('-')[1]}_提现.xlsx"

        wb = Workbook()
        ws = wb.active

        headers = list(items[0].keys())
        ws.append(headers)

        for item in items:
            ws.append(list(item.values()))

        # ⭐ 自动列宽
        for col_idx, col_name in enumerate(headers, 1):
            max_length = len(str(col_name))
            for row in items:
                value = str(row.get(col_name, ""))
                max_length = max(max_length, len(value))

            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        wb.save(fname)

    async def get_all_page(self):
        self.logger.info(f"开始爬取店铺-------------{self.shop_name}------------")
        page = 1
        while True:
            json_data = None
            for attempt in range(3):
                try:
                    cookies = await self.cookie_manager.get_auth()
                    json_data = self.get_info(page, cookies)

                    # ⭐ 核心：统一失效判断
                    if self.is_cookie_invalid(json_data):
                        raise PermissionError("cookie 已失效或接口异常")
                    # 成功直接跳出 retry
                    break

                except PermissionError:
                    self.logger.warning(f"[{self.shop_name}] 登录失效，刷新 cookie（第 {attempt + 1} 次）")
                    await self.cookie_manager.refresh()
                    await asyncio.sleep(2)

                except Exception as e:
                    self.logger.error(f"[{self.shop_name}] 第 {attempt + 1} 次请求失败: {e}")
                    await asyncio.sleep(2)

            if not json_data:
                self.logger.error(f"[{self.shop_name}] 重试失败，终止任务")
                ding_bot_send('me', f"[{self.shop_name}] --financial任务重试失败，终止任务")
                return

            items = self.parse_data(json_data)

            # 数据不为空才进行保存
            if items:
                # 保存数据
                self.save_batch(items)

            self.logger.info(f"店铺-{self.shop_name}-第 {page} 页保存成功")

            # ---------- 没数据，结束 ----------
            if not json_data['info']['list']:
                self.logger.info(f'第{page}页已经没有数据了,程序结束')
                break

            if len(items) < 100:
                self.logger.info("已到最后一页")
                break

            page += 1


# async def run_shein_financial_tixian():
#     month_str = '2025-12'
#     name_list = ["希音全托301-yijia", "希音全托302-juyule", "希音全托303-kedi", "希音全托304-xiyue"]
#     for shop_name in name_list:
#         shein = Shein_Financial_Data_Tixian(shop_name, month_str)
#         await shein.get_all_page()
#
#
# if __name__ == '__main__':
#     asyncio.run(run_shein_financial_tixian())
