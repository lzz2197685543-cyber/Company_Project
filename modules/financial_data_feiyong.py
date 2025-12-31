import requests
from utils.cookie_manager import CookieManager
from utils.logger import get_logger
from pathlib import Path
from datetime import datetime, timezone, timedelta
import asyncio
from utils.dingtalk_bot import ding_bot_send
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

"""shein费用"""

class Shein_Financial_Data_Feiyong:
    def __init__(self,shop_name,month_str):
        self.month_str=month_str
        self.shop_name=shop_name
        self.cookie_manager = CookieManager(shop_name)
        self.cookies = None
        self.headers = {
            'accept-language': 'zh-CN,zh;q=0.9',
            'build-version': '2025-12-25 15:39',
            'origin': 'https://sso.geiwohuo.com',
            'referer': 'https://sso.geiwohuo.com/',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        }
        self.url='https://sso.geiwohuo.com/gsfs/finance/selfReplenish/list'

        self.logger = get_logger("financial_data_feiyong")

    def get_month_date_range(self, month_str: str) -> dict:
        year, month = map(int, month_str.split("-"))

        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year + 1, 1, 1) - timedelta(seconds=1)
        else:
            end = datetime(year, month + 1, 1) - timedelta(seconds=1)

        return {
            "start_date": start.strftime("%Y-%m-%d"),
            "end_date": end.strftime("%Y-%m-%d"),

        }

    # ======================
    # 时间转换
    # ======================

    def _format_time_range(self):
        """
        将 YYYY-MM-DD 转成接口需要的时间格式
        """
        start_time = self.get_month_date_range(self.month_str)["start_date"]
        end_time = self.get_month_date_range(self.month_str)["end_date"]
        start = f"{start_time} 00:00:00"
        end = f"{end_time} 23:59:59"
        return start, end

    def is_cookie_invalid(self, json_data):
        """
        统一判断 cookie 是否失效
        """
        # 请求异常
        if not json_data:
            return True

        # get_info 主动标记
        if json_data.get("msg")=="子系统登录重定向":
            return True

        if not isinstance(json_data, dict):
            return True

        return False

    def get_info(self,page,cookies):
        self.cookies = cookies
        add_time_start, add_time_end = self._format_time_range()
        json_data = {
            'page': page,
            'perPage': 50,
            'tabType': 2,
            'addTimeStart': add_time_start,
            'addTimeEnd': add_time_end,
        }
        try:
            response = requests.post(
                url=self.url,
                cookies=self.cookies,
                headers=self.headers,
                json=json_data,
            )
            print(response.text[:200])
            return response.json()
        except Exception as e:
            self.logger.error(f'请求响应失败:{e}')

    def parse_data(self,json_data):
        try:
            result_list=json_data['info']['data']
        except (KeyError, TypeError):
            self.logger.error("接口数据结构异常")
            return []
        items = []
        for i in result_list:
            items.append({
                "补扣款单号":i.get('replenishNo'),
                "款项类型":i.get('replenishTypeName'),
                "补扣款分类":i.get('categoryName',''),
                "对单类型":i.get('toOrderTypeName',''),
                "关联单据":i.get('relationNo',''),
                "单价":i.get('unitPrice'),
                "数量":i.get('quantity'),
                "总金额":i.get('amount'),
                "币种":i.get('currencyCode'),
                "创建时间":i.get('addTime'),
                "单据状态":i.get('replenishStatusName'),
                "关联报账单":i.get('reportOrderNo'),
                "拒绝原因":i.get('refuseReason',""),
                "确认/拒绝时间":i.get('decisionTime',''),
                "操作人":i.get('operator',''),
                "会计日期":i.get('accountDate',''),
                "是否可报账":i.get('reportableName',''),
                "申诉单号":i.get('appealNo',''),
                "公司主体":i.get('companyName',''),
                "出口模式":i.get('exportingModeName',''),
                "费用类型":i.get('expenseTypeName',''),
                "备注":i.get('remark',''),
            })
        return items

    def save_batch(self, items):
        if not items:
            return

        out_dir = Path(__file__).resolve().parent.parent / "data" / "financial" / "shein"
        out_dir.mkdir(parents=True, exist_ok=True)

        fname = out_dir / f"{self.shop_name}_{datetime.now().strftime('%m')}_费用.xlsx"

        wb = Workbook()
        ws = wb.active

        headers = list(items[0].keys())
        ws.append(headers)

        for item in items:
            ws.append(list(item.values()))

        # ⭐ 自动列宽
        for col_idx, col_name in enumerate(headers, 1):
            max_length = len(str(col_name))
            for row in items:
                value = str(row.get(col_name, ""))
                max_length = max(max_length, len(value))

            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        wb.save(fname)

    async def get_all_page(self):
        self.logger.info(f"开始爬取店铺-------------{self.shop_name}------------")
        page=1
        while True:
            json_data=None
            for attempt in range(3):
                try:
                    cookies = await self.cookie_manager.get_auth()
                    json_data = self.get_info(page, cookies)

                    # ⭐ 核心：统一失效判断
                    if self.is_cookie_invalid(json_data):
                        raise PermissionError("cookie 已失效或接口异常")
                    # 成功直接跳出 retry
                    break
                except PermissionError:
                    self.logger.warning(f"[{self.shop_name}] 登录失效，刷新 cookie（第 {attempt+1} 次）")
                    await self.cookie_manager.refresh()
                    await asyncio.sleep(2)

                except Exception as e:
                    self.logger.error(f"[{self.shop_name}] 第 {attempt+1} 次请求失败: {e}")
                    await asyncio.sleep(2)
            if not json_data:
                self.logger.error(f"[{self.shop_name}] 重试失败，终止任务")
                ding_bot_send('me', f"[{self.shop_name}] --financial任务重试失败，终止任务")
                return

            items = self.parse_data(json_data)

            # 数据不为空才进行保存
            if items:
                # 保存数据
                self.save_batch(items)

            self.logger.info(f"店铺-{self.shop_name}-第 {page} 页保存成功")

            # ---------- 没数据，结束 ----------
            if not json_data['info']['data']:
                self.logger.info(f'第{page}页已经没有数据了,程序结束')
                break

            if len(items) < 100:
                self.logger.info("已到最后一页")
                break

            page += 1


# async def run_shein_financial():
#     name_list = ["希音全托301-yijia", "希音全托302-juyule", "希音全托303-kedi", "希音全托304-xiyue"]
#     month_str='2025-12'
#     for shop_name in name_list:
#         shein = Shein_Financial_Data_Feiyong(shop_name,month_str)
#         await shein.get_all_page()
#
#
# if __name__ == '__main__':
#     asyncio.run(run_shein_financial())

